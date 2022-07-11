#pip install openpyxl

import openpyxl

#CRIAR UMA PLANILHA(BOOK)
book = openpyxl.Workbook()
#COMO VISUALIZAR PAGINAS EXISTENTES
print(book.sheetnames)
#COMO CRIAR UMA PAGINA
book.create_sheet('Frutas')
#COMO SELECIONAR UMA PAGINA
frutas_page = book['Frutas']
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['banana', '5', 'R$3,90'])
frutas_page.append(['maça', '7', 'R$4,50'])
frutas_page.append(['pera', '2', 'R$2,85'])
frutas_page.append(['kiwi', '6', 'R$14,65'])
frutas_page.append(['melao', '2', 'R$13,90'])
#SALVAR A PLANILHA
book.save('Planilha de compras.xlsx')